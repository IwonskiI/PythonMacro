from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import load_workbook
from datetime import datetime
import chromedriver_autoinstaller
import subprocess
import keyboard
import pyperclip
import time


def find_element(xpath):
    global driver
    WebDriverWait(driver, 2).until(
        EC.presence_of_element_located((By.XPATH, xpath)))
    return driver.find_element(By.XPATH, xpath)


def find_element_css(css):
    global driver
    return driver.find_element(By.CSS_SELECTOR, css)


def find_elements_class(className):
    global driver
    return driver.find_elements(By.CSS_SELECTOR, className)


def read_info():
    global storePassword, pluralityStore, clientBrandName
    cnt = 0
    with open("사용법 및 패스워드 입력.txt", "r", encoding='UTF8') as f:
        for line in f:
            cnt += 1
            if cnt == 1:
                storePassword = line.strip().replace("PW: ", "")
            elif cnt == 2:
                pluralityStore = line.strip().replace("다중스토어 여부(Y/N): ", "")
            elif cnt == 3:
                clientBrandName = line.strip().replace("스토어명: ", "")

folder_path = './'
excelName = '상품리스트.xlsx'
sheetName = 'Sheet1'
load_wb = load_workbook(folder_path + excelName, data_only=True)
load_ws = load_wb[sheetName]
subprocess.Popen(
    r'C:\Program Files\Google\Chrome\Application\chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\chrometemp"')  # 디버거 크롬 구동
option = Options()
option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
try:
    driver = webdriver.Chrome(
        f'./{chrome_ver}/chromedriver.exe', options=option)
except:
    chromedriver_autoinstaller.install(True)
    driver = webdriver.Chrome(
        f'./{chrome_ver}/chromedriver.exe', options=option)
smartURL = 'https://sell.smartstore.naver.com/#/login'
loginSolution = 'Seller'
storeID = 'tommyjung415@gmail.com '
read_info()
# storePassword = 'dlrkddls8697!'
# pluralityStore = 'N'
# clientBrandName = 'D:N'
driver.get(smartURL)


def Login():
    global driver, loginSolution, storeID, storePassword
    driver.get(smartURL)
    time.sleep(1)
    if loginSolution != 'Naver':
        find_element(
            '//*[@id="root"]/div/div[1]/div/div/div[4]/div[1]/div/ul[1]/li[1]/input').clear()
        pyperclip.copy(storeID)
        find_element(
            '//*[@id="root"]/div/div[1]/div/div/div[4]/div[1]/div/ul[1]/li[1]/input').send_keys(Keys.CONTROL, 'v')
        find_element(
            '//*[@id="root"]/div/div[1]/div/div/div[4]/div[1]/div/ul[1]/li[2]/input').clear()
        pyperclip.copy(storePassword)
        find_element(
            '//*[@id="root"]/div/div[1]/div/div/div[4]/div[1]/div/ul[1]/li[2]/input').send_keys(Keys.CONTROL, 'v')
        find_element(
            '//*[@id="root"]/div/div[1]/div/div/div[4]/div[1]/div/div/button').click()
    elif loginSolution == 'Naver':
        find_element(
            '//*[@id="root"]/div/div[1]/div/div/div[4]/div[1]/ul/li[2]/button').click()
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[1])
        pyperclip.copy(storeID)
        find_element('//*[@id="id"]').clear()
        find_element('//*[@id="id"]').send_keys(Keys.CONTROL, 'v')
        pyperclip.copy(storePassword)
        find_element('//*[@id="pw"]').clear()
        find_element('//*[@id="pw"]').send_keys(Keys.CONTROL, 'v')
        find_element('//*[@id="log.login"]').click()


def GoToStoreHome():
    global driver, pluralityStore, clientBrandName
    driver.get('https://sell.smartstore.naver.com/#/home/dashboard')
    if pluralityStore == 'Y':
        time.sleep(1)
        find_element(
            '//*[@id="_gnb_nav"]/ul/li[2]/a/i').click()
        time.sleep(1)
        storeName = str(driver.execute_script(
            'return document.getElementsByClassName("text-title")[0].innerText')).replace('스마트스토어', '')
        if storeName == str(clientBrandName):
            find_element(
                '/html/body/div[1]/div/div/div[2]/div/div/div/div/div[1]/div/label/span').click()
        else:
            find_element(
                '/html/body/div[1]/div/div/div[2]/div/div/ul/li/div/div/div/label/span').click()
        time.sleep(1)


# def SetPrice(value):
#     global driver, settingURL, widgetType, saleValue, positionType, my_price, targetURL, delay
#     NaverSetPrice(value)
#     CoupangSetPrice(value)
#     AuctionSetPrice(value)
#     GmarketSetPrice(value)
#     ESMPlusSetPrice(value)
#     ElevenSetPrice(value)
#     TMonSetPrice(value)
#     WMPSetPrice(value)
#     print('가격이 ' + my_price + '원으로 변동되었습니다.')


# def NaverSetPrice(value):
def SetPrice(value):
    global driver, settingURL, widgetType, saleValue, positionType, my_price, targetURL, delay
    driver.get(settingURL)
    time.sleep(2)
    MoveToScroll(
        '//*[@id="productForm"]/ng-include/ui-view[9]/div/div[1]/label')
    time.sleep(1)
    if str(widgetType) == 'y' or str(widgetType) == 'Y':
        find_element('/html/body/div[1]/div/div/div[2]/button/span').click()
    time.sleep(1)
    find_element_css('#prd_price2').clear()
    find_element_css('#prd_price2').send_keys(
        str(int(value) + int(saleValue)))
    MoveToBottop()
    time.sleep(int(delay))
    if int(positionType) == 0:
        find_elements_class('.content')[1].click()
    elif int(positionType) == 1:
        find_element(
            '//*[@id="seller-content"]/ui-view/div[2]/div[2]/button[2]/span[1]').click()
    elif int(positionType) == 2:
        find_elements_class('.content')[3].click()
    time.sleep(1)
    my_price = str(int(value) + int(saleValue))
    time.sleep(2)
    driver.get(targetURL)
    time.sleep(1)


def MoveToScroll(xpath):
    global driver
    tag = find_element(xpath)
    action = ActionChains(driver)
    action.move_to_element(tag).perform()


def MoveToBottop():
    global driver
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')


def SearchOptionName():
    global driver, OptionName
    optionCount = driver.execute_script(
        "return document.getElementsByClassName('filter_on__M8AJ4')[0].parentElement.childElementCount")
    time.sleep(0.5)
    for i in range(int(optionCount)):
        if OptionName == str(find_elements_class('.filter_text__J8EIh')[i].text):
            find_elements_class('.filter_text__J8EIh')[i].click()


def SearchTargetPrice(value):
    global driver, positionType
    time.sleep(0.5)
    if int(value) <= 0:
        if str(positionType) == '0':
            return str(find_element('//*[@id="__next"]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/table/tbody/tr[1]/td[2]/a/em').text.replace(',', ''))
        elif str(positionType) == '1':
            return str(driver.execute_script('return document.getElementsByTagName("tbody")[0].children[0].children[1].children[0].children[0].innerText').replace(',', ''))
    if str(positionType) == '0':
        return str(find_element('//*[@id="__next"]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/table/tbody/tr[' + str(int(value)) + ']/td[2]/a/em').text.replace(',', ''))
    elif str(positionType) == '1':
        return str(driver.execute_script('return document.getElementsByTagName("tbody")[0].children[' + str(int(value) - 1) + '].children[1].children[0].children[0].innerText').replace(',', ''))
    else:
        return str(find_element('//*[@id="__next"]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/table/tbody/tr[' + str(int(value)) + ']/td[2]/a/em').text.replace(',', ''))


def SearchMyPrice():
    global driver, clientBrandName, my_price, rank
    brandCount = driver.execute_script(
        'return document.getElementsByTagName("tbody")[0].childElementCount')
    for i in range(int(brandCount)):
        if clientBrandName == find_elements_class('.productByMall_mall__SIa50')[i].text:
            my_price = SearchTargetPrice(str(i + 1))
            return my_price
    return '999999999999'


def SearchBrandName(value):
    return str(find_elements_class('.productByMall_mall__SIa50')[int(value) - 1].text)


def GetExcelData(i, j):
    global load_ws
    return load_ws.cell(i + 2, j).value


def GetExcelDataLength():
    global load_ws
    for i in range(10000):
        if GetExcelData(i + 1, 1) == None:
            return i+1
    return None

def TimePrint():
    now = datetime.now()
    return str(now.year) + '년 ' + str(now.month) + '월 ' + str(now.day) + '일 ' + str(now.hour) + '시 ' + str(now.minute) + '분 ' + str(now.second) + '초'
    


Login()
print('2차 인증을 완료한 뒤 Ctrl+Shift+1을 눌러주세요.\n(2차 인증창이 안나오는 경우 바로 Ctrl+Shift+1을 눌러주시면 됩니다.)')
print()
print()
while True:
    try:
        if keyboard.is_pressed('ctrl+shift+1'):
            break
    except:
        pass
GoToStoreHome()
loopCount = 0

while True:
    while loopCount < int(GetExcelDataLength()):
        try:
            loopCount += 1
            productName = str(GetExcelData(loopCount-1, 1))
            onOff = str(GetExcelData(loopCount-1, 2))
            targetURL = str(GetExcelData(loopCount-1, 3))
            settingURL = str(GetExcelData(loopCount-1, 4))
            isValidOption = str(GetExcelData(loopCount-1, 5))
            if isValidOption == 'Y' or isValidOption == 'y':
                OptionName = str(GetExcelData(loopCount-1, 6))
            rank = int(GetExcelData(loopCount-1, 7))
            minimumPrice = int(GetExcelData(loopCount-1, 8))
            delay = int(GetExcelData(loopCount-1, 9))
            widgetType = str(GetExcelData(loopCount-1, 10))
            positionType = str(GetExcelData(loopCount-1, 11))
            sniper = str(GetExcelData(loopCount-1, 12))
            myProductURL = str(GetExcelData(loopCount-1, 13))
            targetProductURL = str(GetExcelData(loopCount-1, 14))
            differancePrice = int(GetExcelData(loopCount-1, 15))
            saleValue = int(GetExcelData(loopCount-1, 16))
            if onOff == 'Y' or onOff == 'y':
                if sniper == 'Y' or sniper == 'y':
                    driver.get(myProductURL)
                    time.sleep(1.5)
                    try:
                        print(
                        "상품명 : " + find_element('//*[@id="content"]/div/div[2]/div[2]/fieldset/div[1]/div[1]/h3').text)
                    except:
                        print(str(loopCount + 1) + '번 줄의 상품명을 읽어올 수 없습니다.\n 스나이퍼 부분의 내 URL을 다시 확인 부탁드립니다.\n\n')
                        continue
                    try:
                        my_price = find_element(
                        '//*[@id="content"]/div/div[2]/div[2]/fieldset/div[1]/div[2]/div/strong/span[2]').text.replace(',', '')
                        print('내 가격 : ' + my_price)
                    except:
                        print(str(loopCount + 1) + '번 줄 상품의 가격을 읽어올 수 없습니다.\n 스나이퍼 부분의 내 URL을 다시 확인 부탁드립니다.\n\n')
                        continue
                    driver.get(targetProductURL)
                    time.sleep(1.5)
                    try:
                        target_price = int(find_element(
                        '//*[@id="content"]/div/div[2]/div[2]/fieldset/div[1]/div[2]/div/strong/span[2]').text.replace(',', ''))
                        print('상대 가격 : ' + str(target_price))
                    except:
                        print(str(loopCount + 1) + '번 줄 상품의 가격을 읽어올 수 없습니다.\n 스나이퍼 부분의 상대 URL을 다시 확인 부탁드립니다.\n\n')
                        continue
                    print('설정된 최소 가격 : ' + str(minimumPrice))
                    if minimumPrice >= target_price:
                        if minimumPrice == int(my_price):
                            print('현재 최소 가격으로 설정 되어 있습니다.\n\n')
                            print()
                            time.sleep(delay)
                        else:
                            print('설정된 최소 가격보다 높게 설정되어 있어 가격을 최소로 변경합니다.\n\n')
                            GoToStoreHome()
                            SetPrice(minimumPrice)
                            print(TimePrint()+'에 ' + str(my_price) + '원으로 수정 완료했습니다.')
                            print()
                    else:
                        if target_price - differancePrice != int(my_price):
                            print(
                                '상대 상품의 가격보다 낮게 설정할 수 있어 가격을 상대 상품 가격 - 가격 차이 값으로 변경합니다.\n\n')
                            GoToStoreHome()
                            SetPrice(
                                str(target_price - differancePrice))
                            print(TimePrint()+'에 ' + str(my_price) + '원으로 수정 완료했습니다.')
                            print()
                        else:
                            print('현재 최적의 가격(상대 상품 - 설정한 차이값)으로 설정 되어 있습니다.\n\n')
                            print()
                            time.sleep(delay)
                    driver.get(targetProductURL)
                else:
                    driver.get(targetURL)
                    time.sleep(1)
                    try:
                        product_name = find_element('//*[@id="__next"]/div/div[2]/div[2]/div[1]/h2').text
                        print('상품명 : ' + product_name)
                    except:
                        print(str(loopCount + 1) + '번 줄 상품의 상품명을 읽어올 수 없습니다. 가격비교URL을 다시 확인 부탁드립니다.\n\n')
                        continue    
                    if isValidOption == 'y' or isValidOption == 'Y':
                        try:
                            optionCount = driver.execute_script(
                            'return document.getElementsByClassName("filter_on__M8AJ4")[0].parentElement.childElementCount')
                            if optionCount > 4:
                                MoveToScroll(
                                    '//*[@id="__next"]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/table/thead/tr/th[4]')
                                find_element_css(
                                    '.filter_condition_more__5prbI').click()
                            SearchOptionName()
                        except:
                            print(str(loopCount + 1) + '번 줄 상품의 옵션을 확인해주세요.\n\n')
                            continue
                    try:
                        search_myprice = str(SearchMyPrice())
                        if search_myprice == '999999999999':
                            search_myprice = '현재 상품이 등록되어있지 않습니다.' 
                        print('현재 나의 가격 : ' + search_myprice)
                    except:
                        print(str(loopCount + 1) + '번 줄 상품의 현재 가격을 확인할 수 없습니다. 상품을 다시 확인해주세요.\n\n')
                        continue
                    try:
                        targetrank_brand = str(SearchBrandName(str(rank)))
                        print('현재 설정된 순위의 브랜드 이름 : ' + targetrank_brand)
                    except:
                        print(str(loopCount + 1) + '번 줄 상품의 설정하신 순위에 문제가 있습니다. 순위를 확인해주세요.\nex)순위 범위 초과, 없는 순위 입력 등\n\n')
                        continue
                    print('현재 설정된 순위의 가격 : ' + str(SearchTargetPrice(str(rank))))
                    print('설정된 최소 가격 : ' + str(minimumPrice))
                    if str(clientBrandName) == targetrank_brand:
                        try:
                            nextrank_price = int(SearchTargetPrice(str(rank+1)))
                            if search_myprice == '현재 상품이 등록되어있지 않습니다.':
                                search_myprice = 0
                            else:
                                search_myprice = int(search_myprice)
                            if nextrank_price == search_myprice:
                                if search_myprice == minimumPrice:
                                    print("현재 최적의 가격(최소 가격)으로 설정되어 있습니다.\n\n")
                                    time.sleep(int(delay))
                                else:
                                    if (search_myprice - differancePrice) > minimumPrice:
                                        GoToStoreHome()
                                        SetPrice(str(search_myprice - differancePrice))
                                        print(TimePrint()+'에 ' + str(my_price) + '원으로 수정 완료했습니다.\n\n')
                                    else:
                                        GoToStoreHome()
                                        SetPrice(str(minimumPrice))
                                        print(TimePrint()+'에 ' + str(my_price) + '원으로 수정 완료했습니다.\n\n')
                            else:
                                if nextrank_price - differancePrice > search_myprice:
                                    GoToStoreHome()
                                    SetPrice(str(nextrank_price - differancePrice))
                                    print(TimePrint()+'에 ' + str(my_price) + '원으로 수정 완료했습니다.\n\n')
                                else:
                                    print('현재 최적의 가격(아래 순위 가격 - 설정한 가격 차이)으로 설정되어 있습니다.\n\n')
                                    time.sleep(int(delay))
                        except:
                            print('설정하신 순위가 마지막 순위입니다. 가격을 변경하지 않습니다.')
                            
                    else:
                        target_price = int(SearchTargetPrice(rank))- differancePrice
                        if target_price <= minimumPrice:
                            GoToStoreHome()
                            SetPrice(str(minimumPrice))
                            print(TimePrint()+'에 ' + str(my_price) + '원으로 수정 완료했습니다.\n\n')
                        else:
                            GoToStoreHome()
                            SetPrice(str(target_price))
                            print(TimePrint()+'에 ' + str(my_price) + '원으로 수정 완료했습니다.\n\n')
        except:
            print(str(loopCount + 1) + '번 줄 상품에 문제가 있습니다.\n\n')
    loopCount = 0